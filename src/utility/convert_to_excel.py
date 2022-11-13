import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill, Font, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import sys
import logging
import src.utility.path as util_path

def change_to_correct_dir() -> bool:
    try:
        if getattr(sys, 'frozen', False):
            PATH = os.path.dirname(sys.executable)
            os.chdir(PATH)
            return True
    except:
        return False
    
class Colours:
    colours = {
        "green" : PatternFill("solid", start_color="0000FF00"),
        "blue" : PatternFill("solid", start_color="000000FF"),
        "light green" : PatternFill("solid", start_color="0000FF00"),
        "yellow" : PatternFill("solid", start_color="00FFFF00"),
        "light red" : PatternFill("solid", start_color="00FF0000"),
        "red" : PatternFill("solid", start_color="00FF0000"),
        "light blue" : PatternFill("solid", start_color="000066CC"),
    }
    
class MakeExcel:

    def __init__(self, WORKBOOK_NAME, data_workouts, data_exercises, sheet_theme = "") -> None:
        self.WORKBOOK_NAME = WORKBOOK_NAME
        self.data_workouts = data_workouts
        self.data_exercises = data_exercises
        self.sheet_theme = sheet_theme

        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        
    def make_excel(self) -> bool:
        
        # Move to correct directory if running as exe
        if (util_path.is_running_executable()):
            if not util_path.change_to_correct_dir():
                logging.warning("Couldn't change directory when trying to create excel.")
                return False

        # Sort workouts
        sorted_workouts = self._sort_workout_by_day(self.data_workouts)
        
        # store current day for day matching
        current_day_streak = sorted_workouts[0][2]

        # Starting row/col
        row = 5
        col = 3
        
        # Store first sorted day as current day streak
        self._write_workout_day(row, col, current_day_streak)
        
        # Fill title with colour
        self._fill_header_with_color(row, col)
        
        row_incrementer_for_workout_change = self._find_available_row(row, col) + 1 # + 1 for spacing next workout
        if row_incrementer_for_workout_change == -1 + 1: # -1 represents no safe position and + 1 represents spacing 
            return False
        
        WORKOUT_ROW_INCREMENTER = 2
        default_table_start_cell = "C9"
        default_table_end_cell = "E9"
        TABLE_DEFAULT_ROW = 9
        prev_col = 0
        prev_table_size_ref = ""
        for idx_counter, (w_id, w_name, w_day) in enumerate(sorted_workouts):
            
            # if current day is not the same as before
            if current_day_streak != w_day:
                row = 5
                col += 4
                self._write_workout_day(row, col, w_day)
                self._fill_header_with_color(row, col)
                current_day_streak = w_day
            
            # write workout label
            self._write_workout_label(row + WORKOUT_ROW_INCREMENTER, col)
            
            # write workout name
            self._write_workout_name(row + WORKOUT_ROW_INCREMENTER, col + 1, w_name)
            
            # for exercises with the current workout id
            row_incre = 5
            self._write_exercise_columns(row + row_incre - 1, col)
            
            # write exercises
            row_incre, exercise_count = self._write_exercises_of_workout_id(row, col, row_incre, w_id)
    
            # find ref for table creation
            if exercise_count != 0 and col != prev_col:
                table_size_ref, new_table_char_start, new_table_char_end = self._ref_finder(default_table_start_cell, default_table_end_cell, exercise_count)
                prev_table_size_ref = table_size_ref
                
                default_table_start_cell = f"{new_table_char_start}{TABLE_DEFAULT_ROW}"
                default_table_end_cell = f"{new_table_char_end}{TABLE_DEFAULT_ROW}"
            elif exercise_count != 0 and col == prev_col:
                table_size_ref = prev_table_size_ref
                token_string_array = table_size_ref.split(":")

                table_size_ref = self._table_ref_for_same_day(row_incrementer_for_workout_change, token_string_array)
            else:
                return False
            
            # create table
            self._make_table(f"Table{idx_counter}", table_size_ref)
            
            # increments and new values
            row += row_incrementer_for_workout_change
            if col != prev_col:
                prev_col = col
                
        if self._fix_formatting() == False:
            return False
        
        self.workbook.save(self.WORKBOOK_NAME)
        self.workbook.close()
        return True

    # writes the exercises of a workout id
    # Returns tuple of new row_incre and the amount of exercises
    def _write_exercises_of_workout_id(self, row, col, row_incre, workout_id) -> tuple:
        
        exercise_count = 0
        for e_id, e_name, e_sets, e_reps, e_workout_id in self.data_exercises:
            if e_workout_id == workout_id:
                self.worksheet.cell(row + row_incre, col).value = e_name
                self.worksheet.cell(row + row_incre, col + 1).value = str(e_sets)
                self.worksheet.cell(row + row_incre, col + 2).value = str(e_reps)
                row_incre += 1
                exercise_count += 1

        return (row_incre, exercise_count) 
        
    def _fill_header_with_color(self, row, col):
        WORKOUT_MAX_WIDTH = 3
        
        for i in range(0, WORKOUT_MAX_WIDTH):
            self.worksheet.cell(row, col + i).fill = PatternFill("solid", start_color="FFC000")
    
    def _write_workout_label(self, row, col):
        self.worksheet.cell(row, col).value = "Workout:"
        self.worksheet.cell(row, col).style = "Headline 2"
    
    def _write_workout_day(self, row, col, workout_day):
        self.worksheet.cell(row, col).value = workout_day
        self.worksheet.cell(row, col).style = "Headline 1"
        
    def _write_workout_name(self, row, col, workout_name):
        self.worksheet.cell(row, col).value = workout_name
        self.worksheet.cell(row, col).font = Font(b=True)
        
    def _write_exercise_columns(self, row, col):
        EXERCISE_DATA_COLUMNS = ["Exercises", "Sets", "Reps"]
        for idx, column in enumerate(EXERCISE_DATA_COLUMNS):
            self.worksheet.cell(row, col + idx).value = str(column)
            self.worksheet.cell(row, col + idx).font = Font(color="000000", size=13, b=True)
    
    # returns ref, and the new table chars for next workout
    def _ref_finder(self, first_table_start_cell, first_table_end_cell, exercise_count):
        alphabet = [chr(i) for i in range(65,91)]
        NEXT_WORKOUT_ROW = 4
        
        # if 2 letters
        if first_table_start_cell[1].isalpha():
            table_size_ref = f"{first_table_start_cell}:{first_table_end_cell[0:2]}{int(first_table_end_cell[2:]) + exercise_count + 1}"
        else:
            table_size_ref = f"{first_table_start_cell}:{first_table_end_cell[0]}{int(first_table_end_cell[1:]) + exercise_count + 1}"
        
        # if 2 letters
        if first_table_start_cell[1].isalpha():
            new_table_char_start = "A" + chr(ord(first_table_start_cell[1]) + NEXT_WORKOUT_ROW)
            new_table_char_end = "A" + chr(ord(first_table_end_cell[1]) + NEXT_WORKOUT_ROW)
        else:
            new_table_char_start = chr(ord(first_table_start_cell[0]) + NEXT_WORKOUT_ROW)
            new_table_char_end = chr(ord(first_table_end_cell[0]) + NEXT_WORKOUT_ROW)
        
        # if Z
        if ord(first_table_start_cell[0]) + NEXT_WORKOUT_ROW >= 91:
            new_table_char_start = "A" + alphabet[abs(ord(new_table_char_start) - 91)]
            new_table_char_end = "A" + alphabet[abs(ord(new_table_char_end) - 91)]

        return (table_size_ref, new_table_char_start, new_table_char_end)
    
    def _table_ref_for_same_day(self, row_incrementer_for_workout_change, token_string_array):
        table_ret_to_return = ""
        for i,v in enumerate(token_string_array):
            
            if v[1].isalpha():
                cur_string = v[0:2] + str(int(v[2:]) + row_incrementer_for_workout_change)
            else:
                cur_string = v[0] + str(int(v[1:]) + row_incrementer_for_workout_change)
            
            if i != 1:
                table_ret_to_return += (cur_string) + ":"
            else:
                table_ret_to_return += (cur_string)
                
            cur_string = ""
        return table_ret_to_return
        
    def _make_table(self, name, ref, style_name=""):
        table = Table(displayName=name, ref=ref, headerRowCount=0)
        
        
        if style_name == "":
            style_name = "TableStyleMedium14"
        
        style = TableStyleInfo(name=style_name, showFirstColumn=False,showLastColumn=False, 
                                showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        self.worksheet.add_table(table)

    def _find_available_row(self, row, col) -> bool:
        SAFE_RANGE = 4
        
        is_safe = False
        while not is_safe:
            if (self.worksheet.cell(row, col).value) == None:
                for i in range(SAFE_RANGE):
                    if (self.worksheet.cell(row + i, col).value) != None:
                        break
                    is_safe = True
                    return row
            row += SAFE_RANGE
        
        return -1

    def _sort_workout_by_day(self, workouts_to_sort, sorted_workouts = [], current_day = 0):
        
        MAX_DAY_COUNT = 7
        DAYS = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
        
        # break
        if current_day == MAX_DAY_COUNT:
            return sorted_workouts
        
        # add to final list for current day
        for workout in workouts_to_sort:
            if str(workout[2]).lower() == DAYS[current_day]:
                sorted_workouts.append(workout)
        
        return self._sort_workout_by_day(workouts_to_sort, sorted_workouts, current_day + 1)

    def _cell_styling(self, start_col, end_col, start_row, end_row) -> bool:
        
        # make default colour if no colour is passed
        wanted_cell_colour = PatternFill("solid", start_color="FFFFFF") # white

        # set new style to the colour that is passed
        if self.cell_colour != "":
            for colour, style in Colours.colours.items():
                if colour.lower() == self.cell_colour.lower():
                    wanted_cell_colour = style
                
        # colour cells
        for row in self.worksheet.iter_rows(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col):
            for cell in row:
                col_row = f"{cell.column_letter}{cell.row}"
                self.worksheet[col_row].fill = wanted_cell_colour

    def _fix_formatting(self) -> bool:
        
        dims = {}
        for row in self.worksheet.rows:
            for cell in row:
                if cell.value:
                    col_row = f"{cell.column_letter}{cell.row}"
                    self.worksheet[col_row].alignment = Alignment(horizontal="center")
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        
        for col, value in dims.items():
            self.worksheet.column_dimensions[col].width = value * 1.60
