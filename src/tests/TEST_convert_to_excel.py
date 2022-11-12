import xlsxwriter
import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill, Font, NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import sys
import logging


WORKBOOK = 'output.xlsx'

# Dummy data
hard_coded_workout = [
    (1, "Chest", "Monday"),
    (2, "Back", "Tuesday"),
    (3, "Shoulder", "Wednesday"),
    (5, "Chest & Triceps", "Thursday"),
    (6, "Back & Biceps", "Friday"),
    (7, "Legs & Shoulders", "Saturday"),
    (8, "Abs", "Saturday")
]
# Dummy data
hard_coded_exercises = [
    # id, name, sets, rep range, workout id
    (1, "Bench Press", "4", "8-12", 1),
    (2, "Incline Bench Press", "3", "4-6", 2),
    (3, "Box style", "4", "8-12", 3),
    (4, "Jog1", "3", "4-6", 4),
    (5, "Jog2", "3", "4-6", 5),
    (6, "Jog3", "3", "4-6", 6),
    (7, "Jog4", "3", "4-6", 7),
    (8, "Jog5", "3", "4-6", 8)
]

def change_to_correct_dir() -> bool:
    try:
        if getattr(sys, 'frozen', False):
            PATH = os.path.dirname(sys.executable)
            os.chdir(PATH)
            return True
    except:
        return False
    
class Colours:
    colours = {
        "green" : PatternFill("solid", start_color="0000FF00"),
        "blue" : PatternFill("solid", start_color="000000FF"),
        "light green" : PatternFill("solid", start_color="0000FF00"),
        "yellow" : PatternFill("solid", start_color="00FFFF00"),
        "light red" : PatternFill("solid", start_color="00FF0000"),
        "red" : PatternFill("solid", start_color="00FF0000"),
        "light blue" : PatternFill("solid", start_color="000066CC"),
    }
    
class MakeExcel:

    def __init__(self, WORKBOOK_NAME, data_workouts, data_exercises, cell_colour = "", cell_text_colour = "") -> None:
        self.WORKBOOK_NAME = WORKBOOK_NAME
        self.data_workouts = data_workouts
        self.data_exercises = data_exercises
        self.cell_colour = cell_colour
        self.cell_text_colour = cell_text_colour

        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        
    def make_excel(self) -> bool:
        
        # Move to correct directory if running as exe
        # if (util_path.is_running_executable()):
        #     if not util_path.change_to_correct_dir():
        #         logging.warning("Couldn't change directory when trying to create excel.")
        #         return False

        # Sort workouts
        sorted_workouts = self._sort_workout_by_day(self.data_workouts)
        
        # store current day for day matching
        current_day_streak = sorted_workouts[0][2]

        # Starting row/col
        row = 5
        col = 3
        
        # Store first sorted day as current day streak
        self.worksheet.cell(row, col).value = current_day_streak
        
        self.worksheet.cell(row, col).style = "Headline 1"
        
        # fill heading with colour
        for i in range(0,3):
            self.worksheet.cell(row, col + i).fill = PatternFill("solid", start_color="FFC000")
            
        workout_row_incrementer = 2
        
        row_incrementer_for_workout_change = self._find_available_row(row, col) + 1 # + 1 for spacing next workout
        
        if row_incrementer_for_workout_change == -1 + 1: # -1 represents no safe position and + 1 represents spacing 
            return False
        
        idx_counter = 0
        first_table_start_cell = "C9"
        first_table_end_cell = "E9"
        table_default_row_number = 9
        prev_col = 0
        prev_table_size_ref = ""
        for w_id, w_name, w_day in (sorted_workouts):
            
            # if current day is not the same as before
            if current_day_streak != w_day:
                row = 5
                col += 4
                self.worksheet.cell(row, col).value = w_day
                self.worksheet.cell(row, col).style = "Headline 1"
                table_default_row_number = 9
                # fill heading with colour
                for i in range(0,3):
                    self.worksheet.cell(row, col + i).fill = PatternFill("solid", start_color="FFC000")
                current_day_streak = w_day
            
            # write workout label
            self.worksheet.cell(row + workout_row_incrementer, col).value = "Workout:"
            self.worksheet.cell(row + workout_row_incrementer, col).style = "Headline 2"
            
            # write workout name
            self.worksheet.cell(row + workout_row_incrementer, col + 1).value = w_name
            self.worksheet.cell(row + workout_row_incrementer, col + 1).font = Font(b=True)#, size=16)
            
            # for exercises with the current workout id
            row_incre = 5
            exercise_data_columns = ["Exercises", "Sets", "Reps"]
            for idx, column in enumerate(exercise_data_columns):
                self.worksheet.cell(row + row_incre - 1, col + idx).value = str(column)
                self.worksheet.cell(row + row_incre - 1, col + idx).font = Font(color="FFFFFF", size=13)
            print(f"new column start at {col}")
            exercise_count = 0
            for e_id, e_name, e_sets, e_reps, e_workout_id in self.data_exercises:
                if e_workout_id == w_id:
                    self.worksheet.cell(row + row_incre, col).value = e_name
                    self.worksheet.cell(row + row_incre, col + 1).value = str(e_sets)
                    self.worksheet.cell(row + row_incre, col + 2).value = str(e_reps)
                    row_incre += 1
                    exercise_count += 1
                
            # make table
            if exercise_count != 0 and col != prev_col:

                if first_table_start_cell[1].isalpha():
                    table_size_ref = f"{first_table_start_cell}:{first_table_end_cell[0:2]}{int(first_table_end_cell[2:]) + exercise_count + 1}"
                else:
                    table_size_ref = f"{first_table_start_cell}:{first_table_end_cell[0]}{int(first_table_end_cell[1:]) + exercise_count + 1}"
                
                print(table_size_ref)
                
                alphabet = [chr(i) for i in range(65,91)]
                NEXT_WORKOUT_ROW = 4
                if first_table_start_cell[1].isalpha():
                    new_table_char_start = "A" + chr(ord(first_table_start_cell[1]) + NEXT_WORKOUT_ROW)
                    new_table_char_end = "A" + chr(ord(first_table_end_cell[1]) + NEXT_WORKOUT_ROW)
                else:
                    new_table_char_start = chr(ord(first_table_start_cell[0]) + NEXT_WORKOUT_ROW)
                    new_table_char_end = chr(ord(first_table_end_cell[0]) + NEXT_WORKOUT_ROW)

                if ord(first_table_start_cell[0]) + NEXT_WORKOUT_ROW >= 91:
                    new_table_char_start = "A" + alphabet[abs(ord(new_table_char_start) - 91)]
                    new_table_char_end = "A" + alphabet[abs(ord(new_table_char_end) - 91)]
                
                first_table_start_cell = f"{new_table_char_start}{table_default_row_number}"
                first_table_end_cell = f"{new_table_char_end}{table_default_row_number}"
                prev_table_size_ref = table_size_ref
            else:
                print("in else")
                table_size_ref = prev_table_size_ref
                token_string = table_size_ref.split(":")

                final = ""
                for i,v in enumerate(token_string):
                    
                    if v[1].isalpha():
                        cur_string = v[0:2] + str(int(v[2:]) + row_incrementer_for_workout_change)
                    else:
                        cur_string = v[0] + str(int(v[1:]) + row_incrementer_for_workout_change)
                    
                    if i != 1:
                        final += (cur_string) + ":"
                    else:
                        final += (cur_string)
                        
                    cur_string = ""
                
                table_size_ref = final
                
            self._make_table(f"Table{idx_counter}", table_size_ref)
            
            table_default_row_number = 16
            row += row_incrementer_for_workout_change
            idx_counter += 1
            if col != prev_col:
                prev_col = col
                
        if self._fix_formatting() == False:
            return False
        
        self.workbook.save(self.WORKBOOK_NAME)
        self.workbook.close()
        return True

    def _make_table(self, name, ref, style_name=""):
        table = Table(displayName=name, ref=ref)
        
        if style_name == "":
            style_name = "TableStyleMedium14"
        
        style = TableStyleInfo(name=style_name, showFirstColumn=False,showLastColumn=False, 
                                showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        self.worksheet.add_table(table)

    def _find_available_row(self, row, col) -> bool:
        SAFE_RANGE = 4
        
        is_safe = False
        while not is_safe:
            if (self.worksheet.cell(row, col).value) == None:
                for i in range(SAFE_RANGE):
                    if (self.worksheet.cell(row + i, col).value) != None:
                        break
                    is_safe = True
                    return row
            row += SAFE_RANGE
        
        return -1

    def _sort_workout_by_day(self, workouts_to_sort, sorted_workouts = [], current_day = 0):
        
        MAX_DAY_COUNT = 7
        DAYS = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
        
        # break
        if current_day == MAX_DAY_COUNT:
            return sorted_workouts
        
        # add to final list for current day
        for workout in workouts_to_sort:
            if str(workout[2]).lower() == DAYS[current_day]:
                sorted_workouts.append(workout)
        
        return self._sort_workout_by_day(workouts_to_sort, sorted_workouts, current_day + 1)

    def _cell_styling(self, start_col, end_col, start_row, end_row) -> bool:
        
        # make default colour if no colour is passed
        wanted_cell_colour = PatternFill("solid", start_color="FFFFFF") # white

        # set new style to the colour that is passed
        if self.cell_colour != "":
            for colour, style in Colours.colours.items():
                if colour.lower() == self.cell_colour.lower():
                    wanted_cell_colour = style
                
        # colour cells
        for row in self.worksheet.iter_rows(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col):
            for cell in row:
                col_row = f"{cell.column_letter}{cell.row}"
                self.worksheet[col_row].fill = wanted_cell_colour

    def _fix_formatting(self) -> bool:
        
        dims = {}
        for row in self.worksheet.rows:
            for cell in row:
                if cell.value:
                    col_row = f"{cell.column_letter}{cell.row}"
                    self.worksheet[col_row].alignment = Alignment(horizontal="center")
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        
        for col, value in dims.items():
            self.worksheet.column_dimensions[col].width = value * 1.25

m = MakeExcel(WORKBOOK, hard_coded_workout, hard_coded_exercises)
m.make_excel()