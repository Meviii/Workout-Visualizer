import xlsxwriter
import openpyxl

WORKBOOK = 'output.xlsx'

# data
hard_coded_workout = [
    (1, "Monday", "Chest & Triceps"),
    (2, "Tuesday", "Back & Biceps"),
    (3, "Monday", "Boxing"),
    (4, "Wednesday", "Legs & Shoulders"),
    (5, "Thursday", "Chest & Triceps"),
    (6, "Friday", "Back & Biceps"),
    (7, "Saturday", "Legs & Shoulders"),
    (8, "Saturday", "Abs")
]
# data
hard_coded_exercises = [
    # id, name, sets, rep range, workout id
    (1, "Bench Press", "4", "8-12", 1),
    (2, "Incline Bench Press", "3", "4-6", 1),
    (3, "Box style", "4", "8-12", 3),
    (4, "Jog", "3", "4-6", 3)
]

def make_excel():
    workbook = xlsxwriter.Workbook(rf'{WORKBOOK}')
    worksheet = workbook.add_worksheet()
    workbook.close()
    
    workbook = xlsxwriter.Workbook(WORKBOOK)
    worksheet = workbook.add_worksheet()

    # Starting row/col
    row = 4
    col = 2

    sorted_workouts = sort_workout_by_day(hard_coded_workout)
    
    # store current day for day matching
    current_day_streak = sorted_workouts[0][1] 
    worksheet.write(row, col, current_day_streak) 
    
    workout_row_incrementer = 2
    row_incrementer_for_workout_change = find_available_row(row, col) + 1 # + 1 for spacing next workout
    for w_id, w_day, w_name in (sorted_workouts):
        
        # sort data by day
        # write workout day,
        # write workout name
        # check workout name length, adjust worksheet.set_column_pixels(5, 5, 150)
        # get exercises for that workout using workout id
        # write exercises under workout name
        # if day == same day
        # do not col += 4 and do not re print the day
        
        # if day is same        
        if current_day_streak != w_day:
            row = 4
            col += 4
            worksheet.write(row, col, w_day)
            current_day_streak = w_day
        
        # write workout name
        worksheet.write(row + workout_row_incrementer, col, w_name)
        
        # for exercises with workout id 
        workout_exercises = hard_coded_exercises
        row_incre = 4
        for e_id, e_name, e_sets, e_reps, e_workout_id in workout_exercises:
            if e_workout_id == w_id:
                worksheet.write(row + row_incre, col, e_name)
                worksheet.write(row + row_incre, col + 1, e_sets)
                worksheet.write(row + row_incre, col + 2, e_reps)
                row_incre += 1

        row += row_incrementer_for_workout_change
    
    workbook.close()

def find_available_row(row, col):
    workbook = openpyxl.load_workbook(WORKBOOK)
    worksheet = workbook.active
    SAFE_RANGE = 4
    
    is_safe = False
    while not is_safe:
        if (worksheet.cell(row, col).value) == None:
            for i in range(SAFE_RANGE):
                if (worksheet.cell(row + i, col).value) != None:
                    break
                is_safe = True
                return row
        row += SAFE_RANGE
        
def sort_workout_by_day(workouts_to_sort, sorted_workouts = [], current_day = 0):
    
    MAX_DAY_COUNT = 7
    DAYS = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    
    # break
    if current_day == MAX_DAY_COUNT:
        return sorted_workouts
    
    # add to final list for current day
    for workout in workouts_to_sort:
        if str(workout[1]).lower() == DAYS[current_day]:
            sorted_workouts.append(workout)
    
    return sort_workout_by_day(workouts_to_sort, sorted_workouts, current_day + 1)

def excel_string_width(str):
    """
    Calculate the length of the string in Excel character units. This is only
    an example and won't give accurate results. It will need to be replaced
    by something more rigorous.

    """
    string_width = len(str)

    if string_width == 0:
        return 0
    else:
        return string_width * 1.25
    
def write_value(worksheet,row, col, string):
    min_width = 0

    # Check if it the string is the largest we have seen for this column.
    string_width = excel_string_width(string)
    if string_width > min_width:
        max_width = worksheet.default_col_width
        worksheet.set_column(col, col, width= string_width)
        if string_width > max_width:
            worksheet.set_column(col, col, width = string_width)
    
    worksheet.write(row, col, string)

def _workout_string_tokeniser(string):
    
    mid = 0
    for i in string.split():
        if i.lower() == "on":
            break
        mid += 1
    
    string_builder = ""
    name = string.split()[:mid]
    day = (string.split()[mid+1:])
    for i,v in enumerate(name):
        if i == len(name) - 1:
            string_builder += v    
        else:
            string_builder += v + " "
    name = string_builder
    
    string_builder = ""
    for i,v in enumerate(day):
            string_builder += v
    day = string_builder
    
    return (name, day) 

print(_workout_string_tokeniser("Triceps & Biceps on Monday"))