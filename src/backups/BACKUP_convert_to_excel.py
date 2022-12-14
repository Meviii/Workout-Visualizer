import xlsxwriter
import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill, Font
import os
import sys
import logging
import src.utility.path as util_path

def change_to_correct_dir() -> bool:
    try:
        if getattr(sys, 'frozen', False):
            PATH = os.path.dirname(sys.executable)
            os.chdir(PATH)
            return True
    except:
        return False
    
class Colours:
    colours = {
        "green" : PatternFill("solid", start_color="0000FF00"),
        "blue" : PatternFill("solid", start_color="000000FF"),
        "light green" : PatternFill("solid", start_color="0000FF00"),
        "yellow" : PatternFill("solid", start_color="00FFFF00"),
        "light red" : PatternFill("solid", start_color="00FF0000"),
        "red" : PatternFill("solid", start_color="00FF0000"),
        "light blue" : PatternFill("solid", start_color="000066CC"),
    }
        
class MakeExcel:

    def __init__(self, WORKBOOK_NAME, data_workouts, data_exercises, cell_colour = "", cell_text_colour = "") -> None:
        self.WORKBOOK_NAME = WORKBOOK_NAME
        self.data_workouts = data_workouts
        self.data_exercises = data_exercises
        self.cell_colour = cell_colour
        self.cell_text_colour = cell_text_colour

        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        
    def make_excel(self) -> bool:
        
        # Move to correct directory if running as exe
        if (util_path.is_running_executable()):
            if not util_path.change_to_correct_dir():
                logging.warning("Couldn't change directory when trying to create excel.")
                return False

        # Sort workouts
        sorted_workouts = self._sort_workout_by_day(self.data_workouts)
        
        # store current day for day matching
        current_day_streak = sorted_workouts[0][2]

        # Starting row/col
        row = 5
        col = 3
        
        # Store first sorted day as current day streak
        self.worksheet.cell(row, col).value = current_day_streak

        workout_row_incrementer = 2
        
        row_incrementer_for_workout_change = self._find_available_row(row, col) + 1 # + 1 for spacing next workout
        
        if row_incrementer_for_workout_change == -1 + 1: # -1 represents no safe position and + 1 represents spacing 
            return False
        
        for w_id, w_name, w_day in (sorted_workouts):
            
            self._cell_styling(col - 1, col + 3, row-1, row-1 + row_incrementer_for_workout_change)
            
            # if current day is not the same as before
            if current_day_streak != w_day:
                # self._cell_styling(col, col + 4, row, row + row_incrementer_for_workout_change)
                row = 5
                col += 4
                # print(f"{row} to {row} + {row_incrementer_for_workout_change}")
                self.worksheet.cell(row, col).value = w_day
                current_day_streak = w_day
            
            # write workout name
            self.worksheet.cell(row + workout_row_incrementer, col).value = w_name
            
            # for exercises with workout id
            row_incre = 4
            for e_id, e_name, e_sets, e_reps, e_workout_id in self.data_exercises:
                if e_workout_id == w_id:
                    self.worksheet.cell(row + row_incre, col).value = e_name
                    self.worksheet.cell(row + row_incre, col + 1).value = (e_sets + " sets")
                    self.worksheet.cell(row + row_incre, col + 2).value = (e_reps + " reps")
                    row_incre += 1

            row += row_incrementer_for_workout_change
        
        if self._fix_formatting() == False:
            return False
        
        self.workbook.save(self.WORKBOOK_NAME)
        self.workbook.close()
        return True

    def _find_available_row(self, row, col) -> bool:
        SAFE_RANGE = 4
        
        is_safe = False
        while not is_safe:
            if (self.worksheet.cell(row, col).value) == None:
                for i in range(SAFE_RANGE):
                    if (self.worksheet.cell(row + i, col).value) != None:
                        break
                    is_safe = True
                    return row
            row += SAFE_RANGE
        
        return -1

    def _sort_workout_by_day(self, workouts_to_sort, sorted_workouts = [], current_day = 0):
        
        MAX_DAY_COUNT = 7
        DAYS = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
        
        # break
        if current_day == MAX_DAY_COUNT:
            return sorted_workouts
        
        # add to final list for current day
        for workout in workouts_to_sort:
            if str(workout[2]).lower() == DAYS[current_day]:
                sorted_workouts.append(workout)
        
        return self._sort_workout_by_day(workouts_to_sort, sorted_workouts, current_day + 1)

    def _cell_styling(self, start_col, end_col, start_row, end_row) -> bool:
        
        # make default colour if no colour is passed
        wanted_cell_colour = PatternFill("solid", start_color="FFFFFF") # white

        # set new style to the colour that is passed
        if self.cell_colour != "":
            for colour, style in Colours.colours.items():
                if colour.lower() == self.cell_colour.lower():
                    wanted_cell_colour = style
                
        # colour cells
        for row in self.worksheet.iter_rows(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col):
            for cell in row:
                col_row = f"{cell.column_letter}{cell.row}"
                self.worksheet[col_row].fill = wanted_cell_colour

    def _fix_formatting(self) -> bool:
        
        dims = {}
        for row in self.worksheet.rows:
            for cell in row:
                if cell.value:
                    col_row = f"{cell.column_letter}{cell.row}"
                    self.worksheet[col_row].alignment = Alignment(horizontal="center")
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        
        for col, value in dims.items():
            self.worksheet.column_dimensions[col].width = value * 1.15