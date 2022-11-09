import xlsxwriter
import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill, Font
import os
import sys
import logging
import src.utility.path as util_path

colors = {
    "green" : PatternFill("solid", start_color="0000FF00"),
    "blue" : PatternFill("solid", start_color="000000FF"),
    "light green" : PatternFill("solid", start_color="0000FF00"),
    "yellow" : PatternFill("solid", start_color="00FFFF00"),
    "light red" : PatternFill("solid", start_color="00FF0000"),
    "red" : PatternFill("solid", start_color="00FF0000"),
    "light blue" : PatternFill("solid", start_color="000066CC"),
}

def change_to_correct_dir() -> bool:
    try:
        if getattr(sys, 'frozen', False):
            PATH = os.path.dirname(sys.executable)
            os.chdir(PATH)
            return True
    except:
        return False

def make_excel(data_workouts, data_exercises, WORKBOOK, cell_color="") -> bool:
    
    if (util_path.is_running_executable()):
        if not util_path.change_to_correct_dir():
            logging.warning("Couldn't change directory when trying to create excel.")
            return False
    
    # create file
    try:
        workbook = xlsxwriter.Workbook(WORKBOOK)
        workbook.close()
        
        # open
        workbook = xlsxwriter.Workbook(WORKBOOK)
        worksheet = workbook.add_worksheet()
    except:
        return False
    
    # Starting row/col
    row = 4
    col = 2

    # Sort workouts
    sorted_workouts = sort_workout_by_day(data_workouts)
    
    # store current day for day matching
    current_day_streak = sorted_workouts[0][2]

    worksheet.write(row, col, current_day_streak)
    
    workout_row_incrementer = 2
    row_incrementer_for_workout_change = find_available_row(row, col, WORKBOOK) + 1 # + 1 for spacing next workout
    
    if row_incrementer_for_workout_change == -1 + 1: # -1 = couldn't find, + 1 for spacing
        return False
    
    for w_id, w_name, w_day in (sorted_workouts):
        
        # if current day is not the same as before
        if current_day_streak != w_day:
            row = 4
            col += 4
            worksheet.write(row, col, w_day)
            current_day_streak = w_day
        
        # write workout name
        worksheet.write(row + workout_row_incrementer, col, w_name)
        
        # for exercises with workout id
        row_incre = 4
        for e_id, e_name, e_sets, e_reps, e_workout_id in data_exercises:
            if e_workout_id == w_id:
                worksheet.write(row + row_incre, col, e_name)
                worksheet.write(row + row_incre, col + 1, (e_sets + " sets"))
                worksheet.write(row + row_incre, col + 2, (e_reps + " reps"))
                row_incre += 1

        row += row_incrementer_for_workout_change
        
    workbook.close()
    
    if fix_formatting(WORKBOOK) == False:
        return False

def find_available_row(row, col, WORKBOOK) -> bool:
    try:
        workbook = openpyxl.load_workbook(WORKBOOK)
    except:
        return False
    
    worksheet = workbook.active
    
    SAFE_RANGE = 4
    
    is_safe = False
    while not is_safe:
        if (worksheet.cell(row, col).value) == None:
            for i in range(SAFE_RANGE):
                if (worksheet.cell(row + i, col).value) != None:
                    break
                is_safe = True
                return row
        row += SAFE_RANGE
    
    return -1

def sort_workout_by_day(workouts_to_sort, sorted_workouts = [], current_day = 0):
    
    MAX_DAY_COUNT = 7
    DAYS = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    
    # break
    if current_day == MAX_DAY_COUNT:
        return sorted_workouts
    
    # add to final list for current day
    for workout in workouts_to_sort:
        if str(workout[2]).lower() == DAYS[current_day]:
            sorted_workouts.append(workout)
    
    return sort_workout_by_day(workouts_to_sort, sorted_workouts, current_day + 1)

def cell_styling(WORKBOOK, start_col, end_col, start_row, end_row, cell_color="", cell_text_color="") -> bool:
    try:
        workbook = openpyxl.load_workbook(WORKBOOK)
        worksheet = workbook.active
    except:
        return False
    
    # make default color if no color is passed
    wanted_cell_color = PatternFill("solid", start_color="FFFFFF") # white
    
    # set new style to the color that is passed
    if cell_color != "":
        for color, style in colors:
            if color.lower() == cell_color.lower():
                wanted_cell_color = style
            
    # color cells
    for row in worksheet.iter_rows(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col):
        for cell in row:
            col_row = f"{cell.column_letter}{cell.row}"
            worksheet[col_row].fill = wanted_cell_color
 
    workbook.save(WORKBOOK)

def fix_formatting(WORKBOOK) -> bool:
    try:
        workbook = openpyxl.load_workbook(WORKBOOK)
        worksheet = workbook.active
    except:
        return False
    
    dims = {}
    for row in worksheet.rows:
        for cell in row:
            if cell.value:
                col_row = f"{cell.column_letter}{cell.row}"
                worksheet[col_row].alignment = Alignment(horizontal="center")
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
     
    for col, value in dims.items():
        worksheet.column_dimensions[col].width = value * 1.15
    workbook.save(WORKBOOK)